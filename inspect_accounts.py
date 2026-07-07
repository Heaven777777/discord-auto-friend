import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
fp = os.path.join(BASE_DIR, "账号.xlsx")

lines = []
lines.append(f"文件路径: {fp}")
lines.append(f"存在: {os.path.exists(fp)}")

if not os.path.exists(fp):
    lines.append("文件不存在!")
else:
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    lines.append(f"max_row={ws.max_row}, max_col={ws.max_column}")
    for r in range(1, min(6, ws.max_row) + 1):
        row = [str(ws.cell(row=r, column=c).value) for c in range(1, min(8, ws.max_column) + 1)]
        lines.append(f"Row {r}: {row}")
    wb.close()

result = "\n".join(lines)
print(result)
# 同时写入文件
os.makedirs(os.path.join(BASE_DIR, "logs"), exist_ok=True)
with open(os.path.join(BASE_DIR, "logs", "_excel_preview.txt"), "w", encoding="utf-8") as f:
    f.write(result)