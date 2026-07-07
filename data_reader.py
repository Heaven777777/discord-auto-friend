import os

import openpyxl

import config
import logger


def load_accounts():
    """从 Excel 或文本文件读取账号列表"""
    if not os.path.exists(config.ACCOUNTS_FILE):
        logger.error(f"账号文件不存在: {config.ACCOUNTS_FILE}")
        return []

    accounts = []

    # 先尝试用 openpyxl 读取 .xlsx 格式
    try:
        wb = openpyxl.load_workbook(config.ACCOUNTS_FILE, data_only=True)
        ws = wb.active
        for row_idx in range(config.EXCEL_START_ROW, ws.max_row + 1):
            email = ws.cell(row=row_idx, column=config.EXCEL_EMAIL_COL).value
            password = ws.cell(row=row_idx, column=config.EXCEL_PASSWORD_COL).value
            if not email or not password:
                logger.debug(f"第 {row_idx} 行数据不完整，跳过")
                continue
            accounts.append({
                "row": row_idx,
                "email": str(email).strip(),
                "password": str(password).strip(),
            })
        wb.close()
        logger.info(f"从 Excel 读取到 {len(accounts)} 个账号")
        return accounts
    except Exception:
        logger.debug("openpyxl 无法打开，尝试文本格式读取...")

    # 回退：当作纯文本文件读取（支持制表符/空格分隔）
    try:
        with open(config.ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            row_idx = 1
            for line in f:
                line = line.strip()
                if not line:
                    row_idx += 1
                    continue
                # 支持制表符或多个空格分隔
                parts = line.split("\t") if "\t" in line else line.split()
                if len(parts) >= 2:
                    accounts.append({
                        "row": row_idx,
                        "email": parts[0].strip(),
                        "password": parts[1].strip(),
                    })
                else:
                    logger.debug(f"第 {row_idx} 行格式不正确，跳过: {line}")
                row_idx += 1
        logger.info(f"从文本文件读取到 {len(accounts)} 个账号")
        return accounts
    except Exception as e:
        logger.error(f"无法读取账号文件: {e}")
        return []


def load_friends():
    """从 txt 文件读取好友用户名列表，每行一个，自动排除已添加过的"""
    if not os.path.exists(config.FRIENDS_FILE):
        logger.error(f"好友列表文件不存在: {config.FRIENDS_FILE}")
        return []

    # 加载已添加好友名单
    added_friends = set()
    added_file = os.path.join(config.BASE_DIR, "已添加好友.txt")
    if os.path.exists(added_file):
        with open(added_file, "r", encoding="utf-8") as f:
            for line in f:
                name = line.strip()
                if name:
                    added_friends.add(name)

    friends = []
    skipped = 0
    with open(config.FRIENDS_FILE, "r", encoding="utf-8") as f:
        for line in f:
            name = line.strip()
            if name and not name.startswith("#"):
                if name in added_friends:
                    skipped += 1
                    continue
                friends.append(name)

    if skipped > 0:
        logger.info(f"已跳过 {skipped} 个已添加的好友，剩余 {len(friends)} 个待添加")

    logger.info(f"读取到 {len(friends)} 个目标好友")
    return friends