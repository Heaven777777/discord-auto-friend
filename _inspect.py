import openpyxl
import os
import sys

fp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "账号.xlsx")
out = []

out.append(f"文件路径: {fp}")
out.append(f"存在: {os.path.exists(fp)}")

if not os.path.exists(fp):
    out.append("文件不存在!")
else:
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    out.append(f"max_row={ws.max_row}, max_col={ws.max_column}")
    for r in range(1, min(6, ws.max_row) + 1):
        row = [str(ws.cell(row=r, column=c).value) for c in range(1, min(8, ws.max_column) + 1)]
        out.append(f"Row {r}: {row}")
    wb.close()

result = "\n".join(out)
print(result)
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "_inspect_result.txt"), "w", encoding="utf-8") as f:
    f.write(result)